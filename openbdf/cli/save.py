from pathlib import Path

import pandas as pd
from lib.converters.dataframe import bom_to_dataframe, project_to_dataframe
from lib.types import get_display_type, unwrap_optional
from model.general_info import ProjectRecordBase
from model.tables import BuildingBillMaterialsSlimRecord, ProjectRecord
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill

ATTRIBUTES_NAME_STYLE = NamedStyle(
    "AttributeName", font=Font(bold=True, color="FFFFFF"), fill=PatternFill("solid", fgColor="2b5b47")
)
INFO_STYLE = NamedStyle("Info", font=Font(italic=True))
FIELD_CODE_STYLE = NamedStyle("FieldCode", font=Font(color="FFFFFF"), fill=PatternFill("solid", fgColor="000000"))
REQUIRED_STYLE = NamedStyle("Required", fill=PatternFill("solid", fgColor="a3d3bd"))
RECOMMENDED_STYLE = NamedStyle("Recommended", fill=PatternFill("solid", fgColor="a9d2f6"))
OPTIONAL_STYLE = NamedStyle("Optional", fill=PatternFill("solid", fgColor="d8e6e2"))
ROW_HEADER_STYLE = NamedStyle("RowHeader", font=Font(bold=True), alignment=Alignment(vertical="top"))


def save_openbdf_to_xlsx(
    xlsx_path: str | Path,
    project: ProjectRecord,
    info_sheet_name: str = "General Info",
    bom_sheet_name: str = "Building Bill of Materials",
):
    """Save a list of BuildingBillMaterialsSlimRecord to an Excel file."""

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        info_to_xlsx(writer, project, info_sheet_name)
        bom_to_xlsx(writer, project.bill_of_materials, bom_sheet_name)


def info_to_xlsx(writer: pd.ExcelWriter, project: ProjectRecord, sheet_name: str):
    info_sheet = writer.book.create_sheet(sheet_name)

    project_record_fields = ProjectRecordBase.model_fields  # Base is on purpose to avoid DB-only fields like ID, etc...

    for cell_row, field_code in enumerate(project_record_fields, start=2):  # xlsx are 1-indexed (ewww)
        metadata = project_record_fields.get(field_code)
        extra = metadata.json_schema_extra if metadata else {}
        data_type, is_optional = unwrap_optional(metadata.annotation)
        requirements = extra.get("requirements", "Optional" if is_optional else "Required")

        header_cells = []

        g = info_sheet.cell(row=cell_row, column=1, value=extra.get("group", field_code))
        header_cells.append(g)

        sg = info_sheet.cell(row=cell_row, column=2, value=extra.get("sub_group", field_code))
        header_cells.append(sg)

        an = info_sheet.cell(row=cell_row, column=3, value=extra.get("attribute_name", field_code))
        header_cells.append(an)

        fc = info_sheet.cell(row=cell_row, column=4, value=field_code)
        header_cells.append(fc)

        desc = info_sheet.cell(
            row=cell_row,
            column=5,
            value=metadata.description if metadata and metadata.description else field_code,
        )
        header_cells.append(desc)

        dt = info_sheet.cell(row=cell_row, column=6, value=get_display_type(data_type))
        dt.style = INFO_STYLE

        c = info_sheet.cell(row=cell_row, column=7, value=extra.get("constraint", "none"))
        c.style = INFO_STYLE

        u = info_sheet.cell(row=cell_row, column=8, value=extra.get("units", "none"))
        u.style = INFO_STYLE

        r = info_sheet.cell(row=cell_row, column=9, value=requirements)
        if requirements == "Required":
            r.style = REQUIRED_STYLE
        elif requirements == "Recommended":
            r.style = RECOMMENDED_STYLE
        else:
            r.style = OPTIONAL_STYLE

        info_sheet.cell(row=cell_row, column=10, value=getattr(project, field_code))

        if requirements == "Required":
            for c in header_cells:
                c.style = REQUIRED_STYLE

    a = info_sheet.cell(row=1, column=1, value="Group")
    a.style = ATTRIBUTES_NAME_STYLE
    d = info_sheet.cell(row=1, column=2, value="Sub-group")
    d.style = ATTRIBUTES_NAME_STYLE
    dt = info_sheet.cell(row=1, column=3, value="Attribute Name")
    dt.style = ATTRIBUTES_NAME_STYLE
    c = info_sheet.cell(row=1, column=4, value="Field Code")
    c.style = ATTRIBUTES_NAME_STYLE
    u = info_sheet.cell(row=1, column=5, value="Description")
    u.style = ATTRIBUTES_NAME_STYLE
    r = info_sheet.cell(row=1, column=6, value="Datatype")
    r.style = ATTRIBUTES_NAME_STYLE
    ui = info_sheet.cell(row=1, column=7, value="Constraint")
    ui.style = ATTRIBUTES_NAME_STYLE
    fc = info_sheet.cell(row=1, column=8, value="Units")
    fc.style = ATTRIBUTES_NAME_STYLE
    fc = info_sheet.cell(row=1, column=9, value="Requirements")
    fc.style = ATTRIBUTES_NAME_STYLE
    fc = info_sheet.cell(row=1, column=10, value="Value")
    fc.style = ATTRIBUTES_NAME_STYLE

    info_sheet.column_dimensions["A"].width = 10
    info_sheet.column_dimensions["B"].width = 20
    info_sheet.column_dimensions["C"].width = 30
    info_sheet.column_dimensions["D"].width = 30
    info_sheet.column_dimensions["E"].width = 40
    info_sheet.column_dimensions["F"].width = 30
    info_sheet.column_dimensions["G"].width = 25
    info_sheet.column_dimensions["H"].width = 15
    info_sheet.column_dimensions["I"].width = 15
    info_sheet.column_dimensions["J"].width = 25


def bom_to_xlsx(writer: pd.ExcelWriter, records: list[BuildingBillMaterialsSlimRecord], sheet_name: str):
    df = bom_to_dataframe(records)

    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=1, startrow=8)

    bom_sheet = writer.sheets[sheet_name]
    bom_fields = BuildingBillMaterialsSlimRecord.model_fields

    # headers
    for cell_column, field_code in enumerate(df.columns.tolist(), start=2):  # xlsx are 1-indexed (ewww)
        metadata = bom_fields.get(field_code)
        extra = metadata.json_schema_extra if metadata else {}
        data_type, is_optional = unwrap_optional(metadata.annotation)
        requirements = extra.get("requirements", "Optional" if is_optional else "Required")

        an = bom_sheet.cell(row=1, column=cell_column, value=extra.get("attribute_name", field_code))
        an.style = ATTRIBUTES_NAME_STYLE

        desc = bom_sheet.cell(
            row=2,
            column=cell_column,
            value=metadata.description if metadata and metadata.description else field_code,
        )
        if not is_optional:
            desc.style = REQUIRED_STYLE
        desc.alignment = Alignment(wrap_text=True, vertical="top")

        bom_sheet.cell(row=3, column=cell_column, value=get_display_type(data_type))

        c = bom_sheet.cell(row=4, column=cell_column, value=extra.get("constraint", "none"))
        c.style = INFO_STYLE

        u = bom_sheet.cell(row=5, column=cell_column, value=extra.get("units", "none"))
        u.style = INFO_STYLE

        r = bom_sheet.cell(row=6, column=cell_column, value=requirements)
        if requirements == "Required":
            r.style = REQUIRED_STYLE
        elif requirements == "Recommended":
            r.style = RECOMMENDED_STYLE
        else:
            r.style = OPTIONAL_STYLE

        fc = bom_sheet.cell(row=8, column=cell_column, value=field_code)
        fc.style = FIELD_CODE_STYLE

    a = bom_sheet.cell(row=1, column=1, value="Attribute")
    a.style = ROW_HEADER_STYLE
    d = bom_sheet.cell(row=2, column=1, value="Description")
    d.style = ROW_HEADER_STYLE
    dt = bom_sheet.cell(row=3, column=1, value="Data Type")
    dt.style = ROW_HEADER_STYLE
    c = bom_sheet.cell(row=4, column=1, value="Constraint")
    c.style = ROW_HEADER_STYLE
    u = bom_sheet.cell(row=5, column=1, value="Units")
    u.style = ROW_HEADER_STYLE
    r = bom_sheet.cell(row=6, column=1, value="Requirements")
    r.style = ROW_HEADER_STYLE
    ui = bom_sheet.cell(row=7, column=1, value="User Instructions")
    ui.style = ROW_HEADER_STYLE
    fc = bom_sheet.cell(row=8, column=1, value="Field Code")
    fc.style = ROW_HEADER_STYLE

    bom_sheet.column_dimensions["A"].width = 15
    bom_sheet.column_dimensions["B"].width = 15
    bom_sheet.column_dimensions["C"].width = 20
    bom_sheet.column_dimensions["D"].width = 25
    bom_sheet.column_dimensions["E"].width = 25
    bom_sheet.column_dimensions["F"].width = 25
    bom_sheet.column_dimensions["G"].width = 25
    bom_sheet.column_dimensions["H"].width = 25
    bom_sheet.column_dimensions["I"].width = 35
    bom_sheet.column_dimensions["J"].width = 25
    bom_sheet.column_dimensions["K"].width = 35
    bom_sheet.column_dimensions["L"].width = 20
    bom_sheet.column_dimensions["M"].width = 30
    bom_sheet.column_dimensions["N"].width = 25
    bom_sheet.column_dimensions["O"].width = 25
    bom_sheet.column_dimensions["P"].width = 25
    bom_sheet.column_dimensions["Q"].width = 25
    bom_sheet.column_dimensions["R"].width = 25
    bom_sheet.column_dimensions["S"].width = 25
    bom_sheet.column_dimensions["T"].width = 25
    bom_sheet.column_dimensions["U"].width = 25

    bom_sheet.row_dimensions[2].height = 15
